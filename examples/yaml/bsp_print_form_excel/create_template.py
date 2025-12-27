#!/usr/bin/env python3
"""
Create Excel template for Product Card print form.

This script generates an Excel file with:
- Named Ranges (Заголовок, Данные) for BSP print form areas
- Parameters ({Наименование}, {Код}, etc.) for dynamic values
- Basic formatting (fonts, borders, alignment)

Usage:
    python create_template.py

Output:
    templates/product_card.xlsx

Requirements:
    pip install openpyxl>=3.1.0
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("Error: openpyxl is required")
    print("Install: pip install openpyxl>=3.1.0")
    exit(1)


def create_product_card_template():
    """Create Excel template for product card print form."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Картка"

    # Styles
    title_font = Font(name='Arial', size=16, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    label_font = Font(name='Arial', size=10, bold=True)
    value_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50

    # === Header Area (rows 1-3) ===
    # Row 1: Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "КАРТКА НОМЕНКЛАТУРИ"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Row 2: Empty separator
    ws.row_dimensions[2].height = 10

    # Row 3: Product name (parameter)
    ws.merge_cells('A3:B3')
    ws['A3'] = "{Наименование}"
    ws['A3'].font = header_font
    ws['A3'].alignment = center_align
    ws.row_dimensions[3].height = 25

    # Create Named Range for header (rows 1-3)
    # Note: Named ranges use 1-based indexing and $-anchored references
    wb.defined_names.add(DefinedName("Заголовок", attr_text="Картка!$A$1:$B$3"))

    # === Data Area (rows 4-10) ===
    # Row 4: Empty separator
    ws.row_dimensions[4].height = 10

    # Data fields
    data_rows = [
        ("Код:", "{Код}"),
        ("Артикул:", "{Артикул}"),
        ("Повна назва:", "{НаименованиеПолное}"),
        ("Вид номенклатури:", "{ВидНоменклатуры}"),
        ("Одиниця виміру:", "{ЕдиницаИзмерения}"),
        ("Ставка ПДВ:", "{СтавкаНДС}"),
    ]

    start_row = 5
    for i, (label, value) in enumerate(data_rows):
        row = start_row + i

        # Label cell
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = label_font
        cell_label.alignment = left_align
        cell_label.border = thin_border

        # Value cell (parameter)
        cell_value = ws.cell(row=row, column=2, value=value)
        cell_value.font = value_font
        cell_value.alignment = left_align
        cell_value.border = thin_border

        ws.row_dimensions[row].height = 20

    end_row = start_row + len(data_rows) - 1

    # Create Named Range for data area
    wb.defined_names.add(DefinedName("Данные", attr_text=f"Картка!$A$5:$B${end_row}"))

    # Save
    output_dir = Path(__file__).parent / "templates"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "product_card.xlsx"
    wb.save(output_path)

    print(f"Excel template created: {output_path}")
    print()
    print("Named Ranges:")
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        print(f"  - {name}: {dn.attr_text}")
    print()
    print("Parameters:")
    print("  - {Наименование}")
    print("  - {Код}")
    print("  - {Артикул}")
    print("  - {НаименованиеПолное}")
    print("  - {ВидНоменклатуры}")
    print("  - {ЕдиницаИзмерения}")
    print("  - {СтавкаНДС}")


if __name__ == "__main__":
    create_product_card_template()

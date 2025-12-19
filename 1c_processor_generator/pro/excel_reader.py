   
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    from openpyxl.styles import Font, Alignment, Border, PatternFill
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


                                              
PARAMETER_PATTERN = re.compile(r'\{([A-Za-zА-Яа-яЁёҐґЇїІіЄє_][A-Za-zА-Яа-яЁёҐґЇїІіЄє0-9_]*)\}')


@dataclass
class CellData:
                                                            
    row: int                                       
    col: int                                          
    value: Optional[str] = None
    parameters: List[str] = field(default_factory=list)                         

                
    font_name: str = "Arial"
    font_size: float = 9.0
    font_bold: bool = False
    font_italic: bool = False

    horizontal_alignment: str = "Left"                       
    vertical_alignment: str = "Top"                          

                                  
    border_left: bool = False
    border_right: bool = False
    border_top: bool = False
    border_bottom: bool = False

                                      
    background_color: Optional[str] = None

                   
    wrap_text: bool = False


@dataclass
class MergeRange:
                                         
    start_row: int               
    start_col: int               
    end_row: int                 
    end_col: int                 

    @property
    def width(self) -> int:
                                                               
        return self.end_col - self.start_col


@dataclass
class NamedArea:
                                                                            
    name: str
    start_row: int               
    end_row: int                 
    start_col: int                                 
    end_col: int                                   
    area_type: str = "Rows"                   


@dataclass
class ColumnDef:
                                       
    index: int                           
    width: float                                       


@dataclass
class RowDef:
                                     
    index: int                           
    height: float                         


class ExcelReader:
           

                                                                 
    DEFAULT_COLUMN_WIDTH = 64

                                                        
    CHAR_WIDTH_MULTIPLIER = 7.5

    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
                   
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel reading. "
                "Install it with: pip install openpyxl>=3.1.0"
            )

        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        self.wb = load_workbook(str(self.file_path), data_only=True)

        if sheet_name:
            if sheet_name not in self.wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.wb.sheetnames}")
            self.ws: Worksheet = self.wb[sheet_name]
        else:
            self.ws: Worksheet = self.wb.active

        self._cells_cache: Optional[Dict[Tuple[int, int], CellData]] = None

    def get_dimensions(self) -> Tuple[int, int, int, int]:
                   
        return (
            self.ws.min_row or 1,
            self.ws.min_column or 1,
            self.ws.max_row or 1,
            self.ws.max_column or 1
        )

    def get_cells(self) -> Dict[Tuple[int, int], CellData]:
                   
        if self._cells_cache is not None:
            return self._cells_cache

        cells: Dict[Tuple[int, int], CellData] = {}

        min_row, min_col, max_row, max_col = self.get_dimensions()

        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell_data = self._parse_cell(cell)
                cells[(row_idx, col_idx)] = cell_data

        self._cells_cache = cells
        return cells

    def _parse_cell(self, cell: "Cell") -> CellData:
                                              
        value = str(cell.value) if cell.value is not None else None

                                    
        parameters = []
        if value:
            parameters = PARAMETER_PATTERN.findall(value)

                            
        font = cell.font
        alignment = cell.alignment
        border = cell.border
        fill = cell.fill

                                      
        h_align_map = {
            "left": "Left",
            "center": "Center",
            "right": "Right",
            "general": "Left",
            None: "Left"
        }

                                    
        v_align_map = {
            "top": "Top",
            "center": "Center",
            "bottom": "Bottom",
            None: "Top"
        }

                       
        def has_border(side) -> bool:
            return side is not None and side.style is not None and side.style != "none"

                          
        bg_color = None
        if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
            rgb = fill.fgColor.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                bg_color = rgb[-6:]                                    

        return CellData(
            row=cell.row,
            col=cell.column,
            value=value,
            parameters=parameters,
            font_name=font.name or "Arial",
            font_size=font.size or 9.0,
            font_bold=font.bold or False,
            font_italic=font.italic or False,
            horizontal_alignment=h_align_map.get(alignment.horizontal, "Left") if alignment else "Left",
            vertical_alignment=v_align_map.get(alignment.vertical, "Top") if alignment else "Top",
            border_left=has_border(border.left) if border else False,
            border_right=has_border(border.right) if border else False,
            border_top=has_border(border.top) if border else False,
            border_bottom=has_border(border.bottom) if border else False,
            background_color=bg_color,
            wrap_text=alignment.wrap_text if alignment and alignment.wrap_text else False
        )

    def get_merged_cells(self) -> List[MergeRange]:
                   
        merged = []
        for merge_range in self.ws.merged_cells.ranges:
            merged.append(MergeRange(
                start_row=merge_range.min_row,
                start_col=merge_range.min_col,
                end_row=merge_range.max_row,
                end_col=merge_range.max_col
            ))
        return merged

    def get_named_areas(self) -> List[NamedArea]:
                   
        areas = []

                                                                            
                                                       
        try:
                                        
            defined_names_list = list(self.wb.defined_names.values())
        except AttributeError:
                                                
            defined_names_list = self.wb.defined_names.definedName

        for defined_name in defined_names_list:
            name = defined_name.name

                                       
            if name.startswith("_"):
                continue

                                             
            try:
                for sheet_title, coord in defined_name.destinations:
                    if sheet_title != self.ws.title:
                        continue

                                                                     
                    area = self._parse_range_coordinate(coord, name)
                    if area:
                        areas.append(area)
            except (AttributeError, ValueError):
                continue

        return areas

    def _parse_range_coordinate(self, coord: str, name: str) -> Optional[NamedArea]:
                                                          
                        
        coord = coord.replace("$", "")

        if ":" not in coord:
                                                      
            return None

        start, end = coord.split(":")

                                                                             
        start_row, start_col = self._parse_cell_ref(start)
        end_row, end_col = self._parse_cell_ref(end)

                             
        if start_col == -1 and end_col == -1:
                              
            area_type = "Rows"
        elif start_row == -1 and end_row == -1:
                                 
            area_type = "Columns"
        else:
                                        
            area_type = "Rows"

        return NamedArea(
            name=name,
            start_row=start_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            area_type=area_type
        )

    def _parse_cell_ref(self, ref: str) -> Tuple[int, int]:
                   
                                                       
        col_match = re.match(r'^([A-Za-z]+)', ref)
        row_match = re.search(r'(\d+)$', ref)

        col = -1
        row = -1

        if col_match:
            col = column_index_from_string(col_match.group(1))

        if row_match:
            row = int(row_match.group(1))

        return row, col

    def get_columns(self) -> List[ColumnDef]:
                   
        columns = []
        _, min_col, _, max_col = self.get_dimensions()

        for col_idx in range(min_col, max_col + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = self.ws.column_dimensions.get(col_letter)

            if col_dim and col_dim.width:
                width = col_dim.width * self.CHAR_WIDTH_MULTIPLIER
            else:
                width = self.DEFAULT_COLUMN_WIDTH

            columns.append(ColumnDef(
                index=col_idx - 1,                   
                width=round(width)
            ))

        return columns

    def get_rows(self) -> List[RowDef]:
                   
        rows = []
        min_row, _, max_row, _ = self.get_dimensions()

        for row_idx in range(min_row, max_row + 1):
            row_dim = self.ws.row_dimensions.get(row_idx)

            if row_dim and row_dim.height:
                height = row_dim.height
            else:
                height = 15.0                      

            rows.append(RowDef(
                index=row_idx - 1,                   
                height=height
            ))

        return rows

    def find_parameters(self) -> List[Tuple[int, int, str]]:
                   
        params = []
        cells = self.get_cells()

        for (row, col), cell_data in cells.items():
            for param_name in cell_data.parameters:
                params.append((row, col, param_name))

        return params

    def get_unique_fonts(self) -> List[Dict[str, Any]]:
                   
        fonts = {}
        cells = self.get_cells()

        for cell_data in cells.values():
            font_key = (
                cell_data.font_name,
                cell_data.font_size,
                cell_data.font_bold,
                cell_data.font_italic
            )

            if font_key not in fonts:
                fonts[font_key] = {
                    "name": cell_data.font_name,
                    "size": cell_data.font_size,
                    "bold": cell_data.font_bold,
                    "italic": cell_data.font_italic
                }

        return list(fonts.values())

    def close(self):
                                 
        self.wb.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def check_openpyxl_available() -> bool:
                                         
    return OPENPYXL_AVAILABLE
